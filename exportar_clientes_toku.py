"""
Exporta clientes (RD Station + Omie) para o padrão de importação da Toku.
Somente leitura (SELECT) — as extrações em sql/extract_*.sql não gravam nada
em nenhuma das bases. Todo o tratamento de dados (nome, CPF/CNPJ, e-mail,
telefone, CEP/cidade/UF, deduplicação) é feito em tratamento_clientes_toku.py.

Uso:
    python exportar_clientes_toku.py            # carga completa
    python exportar_clientes_toku.py --teste    # limita Deals a TOP 1000
"""
import argparse
import os
from datetime import datetime

import dotenv
import pandas as pd
import pyodbc
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import tratamento_clientes_toku as tratamento

dotenv.load_dotenv()

SQL_SERVER   = os.getenv("SQL_SERVER", "")
SQL_PORT     = os.getenv("SQL_PORT", "1433")
SQL_DATABASE = os.getenv("SQL_DATABASE", "")
SQL_USER     = os.getenv("SQL_USER", "")
SQL_PASSWORD = os.getenv("SQL_PASSWORD", "")
SQL_DRIVER   = os.getenv("SQL_DRIVER", "{ODBC Driver 18 for SQL Server}")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
SQL_DIR    = os.path.join(BASE_DIR, "sql")
OUTPUT_DIR = os.path.join(BASE_DIR, "exports")

TAMANHO_LOTE = 1000

COLUNAS_TEXTO = {"CPF OU CNPJ (SOMENTE NÚMEROS)", "CEP"}

# Layout padrão do modelo de importação da Toku (aba "Planilha1").
SHEET_NAME = "Planilha1"
HEADER_FONT = Font(name="Aptos Narrow", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_ROW_HEIGHT = 38.5
COLUNA_LARGURAS = {
    "A": 13.1796875,
    "B": 19.81640625,
    "C": 8.90625,
    "D": 32.08984375,
    "E": 22.90625,
    "F": 11.453125,
    "G": 12.0,
    "H": 18.7265625,
}


def validar_variaveis_ambiente():
    obrigatorias = {
        "SQL_SERVER": SQL_SERVER,
        "SQL_DATABASE": SQL_DATABASE,
        "SQL_USER": SQL_USER,
        "SQL_PASSWORD": SQL_PASSWORD,
    }
    faltando = [nome for nome, valor in obrigatorias.items() if not valor]
    if faltando:
        raise SystemExit(
            "Variavel(is) de ambiente vazia(s) ou nao encontrada(s): "
            f"{', '.join(faltando)}. Confira se o arquivo .env esta na raiz do "
            "projeto, salvo (sem alteracoes pendentes) e sem BOM/UTF-8 com "
            "assinatura no inicio do arquivo."
        )


def connect():
    # "tcp:" força o protocolo TCP/IP explicitamente — sem isso, alguns clientes
    # Windows tentam Named Pipes primeiro (configuração de protocolo do cliente)
    # e a conexão falha mesmo com a porta acessível.
    conn_str = (
        f"DRIVER={SQL_DRIVER};SERVER=tcp:{SQL_SERVER},{SQL_PORT};DATABASE={SQL_DATABASE};"
        f"UID={SQL_USER};PWD={SQL_PASSWORD};TrustServerCertificate=yes"
    )
    return pyodbc.connect(conn_str, timeout=30)


def carregar_query(nome_arquivo: str) -> str:
    with open(os.path.join(SQL_DIR, nome_arquivo), "r", encoding="utf-8") as f:
        return f.read()


def consultar_em_lotes(conn, nome_arquivo: str, valores, tamanho_lote: int = TAMANHO_LOTE) -> pd.DataFrame:
    """Roda um extract sql/*.sql que tem um placeholder {IN_LIST}, quebrando
    `valores` em lotes (limite de parâmetros do SQL Server) e concatenando
    o resultado. Substitui os JOINs de narrowing (docs_alvo/emails_alvo/...)
    que existiam na query monolítica antiga."""
    if not valores:
        return pd.DataFrame()
    template = carregar_query(nome_arquivo)
    partes = []
    for inicio in range(0, len(valores), tamanho_lote):
        lote = valores[inicio:inicio + tamanho_lote]
        placeholders = ",".join("?" for _ in lote)
        query = template.replace("{IN_LIST}", placeholders)
        partes.append(pd.read_sql(query, conn, params=lote))
    return pd.concat(partes, ignore_index=True)


def montar_dataframe(conn, teste: bool) -> pd.DataFrame:
    top_n = "TOP 1000" if teste else ""
    query_deals = carregar_query("extract_rd_deals.sql").replace("/*TOP_N*/", top_n)

    print("Extraindo Deals (RD Station)...")
    df_deals = pd.read_sql(query_deals, conn)

    print("Extraindo campos customizados dos Deals (RD Station)...")
    df_custom_fields = consultar_em_lotes(
        conn, "extract_rd_deal_custom_fields.sql", df_deals["DealId"].tolist(),
    )
    base = tratamento.montar_base(df_deals, df_custom_fields)

    print("Extraindo Contacts (RD Station)...")
    df_contacts = pd.read_sql(carregar_query("extract_rd_contacts.sql"), conn)
    contato_uni = tratamento.montar_contato_uni(df_contacts)
    contact_ids = tratamento.contact_ids_para_buscar(contato_uni)

    print("Extraindo telefones de Contacts (RD Station)...")
    df_contact_phones = consultar_em_lotes(conn, "extract_rd_contact_phones.sql", contact_ids)
    fone_uni = tratamento.montar_fone_uni(df_contact_phones)

    print("Extraindo e-mails secundários de Contacts (RD Station)...")
    df_contact_emails = consultar_em_lotes(conn, "extract_rd_contact_emails.sql", contact_ids)
    email_secundario = tratamento.montar_email_secundario(df_contact_emails)

    print("Extraindo cliente_fornecedor (Omie)...")
    df_omie = pd.read_sql(carregar_query("extract_omie_clientes.sql"), conn)
    omie_uni = tratamento.montar_omie_uni(df_omie)
    omie_email_uni = tratamento.montar_omie_email_uni(df_omie)

    print("Extraindo CepEndereco (energyClub)...")
    df_cep = pd.read_sql(carregar_query("extract_energyclub_cep.sql"), conn)
    cep_tab = tratamento.montar_cep_tab(df_cep)

    print("Tratando e montando planilha final...")
    return tratamento.montar_planilha_final(
        base, contato_uni, fone_uni, email_secundario, omie_uni, cep_tab, omie_email_uni,
    )


def main():
    parser = argparse.ArgumentParser(description="Exporta clientes RD Station + Omie para o padrão Toku.")
    parser.add_argument(
        "--teste", action="store_true",
        help="Limita o extract de Deals a TOP 1000, para validar o resultado antes da exportação completa.",
    )
    args = parser.parse_args()

    validar_variaveis_ambiente()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("Conectando ao SQL Server...")
    conn = connect()
    try:
        df = montar_dataframe(conn, teste=args.teste)
    finally:
        conn.close()

    print(f"{len(df)} clientes retornados.")
    campos_completo = ["CELULAR (PADRÃO E.164) Ex. +5548999567788", "CEP", "CIDADE", "CÓDIGO DO ESTADO"]
    if all(col in df.columns for col in campos_completo):
        prontos = int(df[campos_completo].notna().all(axis=1).sum())
        print(f"  - prontos para Toku: {prontos}")
        print(f"  - com alguma pendencia: {len(df) - prontos}")

    print("Nulos por coluna:")
    for col in df.columns:
        nulos = int(df[col].isna().sum())
        if nulos:
            print(f"  - {col}: {nulos}")

    for col in COLUNAS_TEXTO:
        if col in df.columns:
            df[col] = df[col].astype("string")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"clientes_toku_{timestamp}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        ws = writer.sheets[SHEET_NAME]

        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
        for idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(idx)
            ws.column_dimensions[letra].width = COLUNA_LARGURAS.get(letra, 15)
            ws[f"{letra}1"].font = HEADER_FONT
            ws[f"{letra}1"].alignment = HEADER_ALIGNMENT
            if col_name in COLUNAS_TEXTO:
                for cell in ws[letra][1:]:
                    cell.number_format = "@"

    print(f"Arquivo gerado em: {output_path}")


if __name__ == "__main__":
    main()
